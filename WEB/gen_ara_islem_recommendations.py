import os
import json
from openpyxl import load_workbook


def norm(s):
    s = "" if s is None else str(s)
    s = s.lower()
    # Approximate Turkish chars
    s = (
        s.replace("ı", "i")
        .replace("ğ", "g")
        .replace("ü", "u")
        .replace("ş", "s")
        .replace("ö", "o")
        .replace("ç", "c")
    )
    return s


def norm_header(s):
    return norm(s).strip()


def main():
    # Excel kaynakları: proje kökündeki DATA/ (yoksa bir üst dizin)
    script_dir = os.path.dirname(__file__)
    root = os.path.abspath(os.path.join(script_dir, ".."))
    data_dir = os.path.join(root, "DATA")
    source_dir = data_dir if os.path.isdir(data_dir) else root
    os.chdir(source_dir)

    files = [f for f in os.listdir(".") if f.lower().endswith(".xlsx")]
    files.sort()

    loaded = []
    for fn in files:
        wb = load_workbook(fn, read_only=False, data_only=True)
        sh = wb[wb.sheetnames[0]]
        loaded.append((fn, sh.max_row or 0, sh.max_column or 0))

    cost_fn = None
    base_fns = []
    for fn, mr, _mc in loaded:
        if mr > 200 and cost_fn is None:
            cost_fn = fn
        else:
            base_fns.append(fn)

    # code(int) -> base fields from “Ara İşlem Tablosu” and “Düzenli Kılavuz”
    lookup = {}

    for fn in base_fns:
        wb = load_workbook(fn, read_only=False, data_only=True)
        sh = wb[wb.sheetnames[0]]
        max_row = sh.max_row or 0
        max_col = sh.max_column or 0
        header = [sh.cell(1, c).value for c in range(1, max_col + 1)]

        atik_code_idx = None
        desc_idx = None
        status_idx = None
        op_idx = None
        src_idx = None

        for i, h in enumerate(header, start=1):
            nh = norm_header(h)
            if atik_code_idx is None and ("atik" in nh and "kodu" in nh):
                atik_code_idx = i
            if desc_idx is None and "atik tanim" in nh:
                desc_idx = i
            if status_idx is None and "dur" in nh:
                status_idx = i
            if op_idx is None and "uygulanan" in nh:
                op_idx = i
            if src_idx is None and "kaynak" in nh:
                src_idx = i

        # Fallback indices (from your earlier structure analysis)
        if atik_code_idx is None:
            atik_code_idx = 2
        if desc_idx is None:
            desc_idx = 3
        if status_idx is None:
            status_idx = 5
        if op_idx is None:
            op_idx = 12
        if src_idx is None:
            src_idx = 13

        for r in range(2, max_row + 1):
            code = sh.cell(r, atik_code_idx).value
            if code is None:
                continue
            try:
                code_int = int(float(code))
            except Exception:
                continue

            desc = sh.cell(r, desc_idx).value
            status = sh.cell(r, status_idx).value
            op = sh.cell(r, op_idx).value
            src = sh.cell(r, src_idx).value

            cur = lookup.get(code_int)
            if cur is None:
                lookup[code_int] = {
                    "atikTanimi": desc or "",
                    "durumu": status or "",
                    "araIslem": op or "",
                    "kaynak": src or "",
                }
            else:
                if desc and not cur.get("atikTanimi"):
                    cur["atikTanimi"] = desc
                if status and not cur.get("durumu"):
                    cur["durumu"] = status
                if op and (len(str(op)) > len(str(cur.get("araIslem", "")))):
                    cur["araIslem"] = op
                if src and not cur.get("kaynak"):
                    cur["kaynak"] = src

    # “Düzenlenmiş 46 Ara İşlem” for cost + islemTipi
    cost_lookup = {}
    if cost_fn:
        wb = load_workbook(cost_fn, read_only=False, data_only=True)
        sh = wb[wb.sheetnames[0]]
        max_row = sh.max_row or 0
        max_col = sh.max_column or 0
        header = [sh.cell(1, c).value for c in range(1, max_col + 1)]

        atik_code_idx = 2
        desc_idx = 3
        cost_idx = 4
        type_idx = 5
        op_idx = 6
        src_idx = 7

        for i, h in enumerate(header, start=1):
            nh = norm_header(h)
            if "atik" in nh and "kodu" in nh:
                atik_code_idx = i
            if "maliyet" in nh:
                cost_idx = i
            if "islem" in nh and "tip" in nh:
                type_idx = i
            if "uygulanan" in nh:
                op_idx = i
            if "kaynak" in nh:
                src_idx = i
            if "atik tanim" in nh:
                desc_idx = i

        for r in range(2, max_row + 1):
            code = sh.cell(r, atik_code_idx).value
            if code is None:
                continue
            try:
                code_int = int(float(code))
            except Exception:
                continue

            cost = sh.cell(r, cost_idx).value
            tip = sh.cell(r, type_idx).value
            op = sh.cell(r, op_idx).value
            src = sh.cell(r, src_idx).value

            cur = cost_lookup.get(code_int)
            if cur is None:
                cost_lookup[code_int] = {
                    "maliyetPuan": cost,
                    "islemTipi": tip,
                    "araIslem": op or "",
                    "kaynak": src or "",
                }
            else:
                if cost is not None and cur.get("maliyetPuan") is None:
                    cur["maliyetPuan"] = cost
                try:
                    if cost is not None and cur.get("maliyetPuan") is not None:
                        if float(cost) > float(cur["maliyetPuan"]):
                            cur["maliyetPuan"] = cost
                except Exception:
                    pass

                if op and (
                    (not cur.get("araIslem")) or len(str(op)) > len(str(cur.get("araIslem", "")))
                ):
                    cur["araIslem"] = op
                if src and not cur.get("kaynak"):
                    cur["kaynak"] = src

    # merge
    for code, base in list(lookup.items()):
        c = cost_lookup.get(code)
        if not c:
            continue
        base["maliyetPuan"] = c.get("maliyetPuan")
        base["islemTipi"] = c.get("islemTipi")
        if c.get("araIslem") and (
            (not base.get("araIslem")) or len(str(c.get("araIslem"))) > len(str(base.get("araIslem", "")))
        ):
            base["araIslem"] = c.get("araIslem", "")
        if c.get("kaynak") and not base.get("kaynak"):
            base["kaynak"] = c.get("kaynak", "")

    # UI/lookup category keywords.
    # Not: lookup anahtarlarını Türkçe karakter içermeyen ASCII/normalize formda tutuyoruz.
    categories = {
        # blokzincir
        "kagit / karton": {"keywords": ["kagit", "karton", "paper"]},
        "plastik": {"keywords": ["plastik", "pet", "ambalaj"]},
        "cam": {"keywords": ["cam"]},
        "metal": {"keywords": ["metal", "hurda", "kablo", "bakir"]},
        "tehlikeli atik": {"keywords": ["tehlikeli", "kimyasal", "ambalaj"]},
        "organik": {"keywords": ["organik", "biyolojik", "gida", "kompost"]},
        "elektronik": {"keywords": ["elektronik", "weee", "kablo"]},
        # ilanlarım / canlı pazar
        "karton ve ambalaj": {"keywords": ["karton", "kagit", "ambalaj"]},
        "metal ve elektronik": {"keywords": ["metal", "elektronik", "kablo", "hurda"]},
        "karisik": {"keywords": []},
    }

    ui_categories = [
        "kagit / karton",
        "plastik",
        "cam",
        "metal",
        "tehlikeli atik",
        "organik",
        "elektronik",
        "karton ve ambalaj",
        "metal ve elektronik",
        "karisik",
    ]

    chat_categories = {
        "plastik": ["plastik", "pet", "ambalaj"],
        "kagit": ["karton", "kagit", "kağıt"],
        "metal": ["metal", "hurda", "kablo", "bakir"],
        "elektronik": ["elektronik", "weee", "kablo"],
        "tehlikeli": ["tehlikeli", "kimyasal"],
        "organik": ["organik", "biyolojik", "gida"],
    }

    def pick_best_for_category(cat_name: str):
        cat = categories.get(cat_name, {"keywords": []})
        keywords = [norm(k) for k in cat.get("keywords", [])]

        best = None
        best_score = -10**18

        for code, row in lookup.items():
            desc = norm(row.get("atikTanimi"))
            op = norm(row.get("araIslem"))
            status = norm(row.get("durumu"))

            s = 0
            for kw in keywords:
                if not kw:
                    continue
                if kw in desc:
                    s += 2
                if kw in op:
                    s += 1

            # tehlikeli alignment
            if "tehlikeli" in status:
                if "tehlikeli" in norm(cat_name):
                    s += 30
                else:
                    s -= 10
            if "tehlikeli" in norm(cat_name) and "tehlikeli" not in status:
                s -= 20

            # cost bonus
            cost = row.get("maliyetPuan")
            if cost is not None and cost != "":
                try:
                    s += float(cost)
                except Exception:
                    pass

            # islemTipi bonus
            tip = row.get("islemTipi")
            if tip is not None and tip != "":
                try:
                    if str(float(tip)) in ("1.0", "1"):
                        s += 2
                except Exception:
                    pass

            if best is None or s > best_score:
                best_score = s
                best = {"atikKodu": code, **row, "_score": s}

        return best

    recommendations = {}
    for cat_name in ui_categories:
        best = pick_best_for_category(cat_name)
        if not best:
            continue
        recommendations[cat_name] = {
            "atikKodu": best.get("atikKodu"),
            "araIslem": best.get("araIslem", ""),
            "kaynak": best.get("kaynak", ""),
            "maliyetPuan": best.get("maliyetPuan"),
            "islemTipi": best.get("islemTipi"),
            "durumu": best.get("durumu", ""),
            "matchScore": best.get("_score"),
        }

    # Çıktı: WEB/assets (sayfa bu dosyayı yükler)
    out_paths = [
        os.path.join(script_dir, "assets", "ara-islem-recommendations.js"),
    ]

    rec_json = json.dumps(recommendations, ensure_ascii=True, separators=(",", ":"))
    chat_json = json.dumps(chat_categories, ensure_ascii=True, separators=(",", ":"))

    payload_lines = []
    payload_lines.append("(function(){\n")
    payload_lines.append("  window.ARA_ISLEM_RECOMMENDATIONS = " + rec_json + ";\n")
    payload_lines.append("  window.ARA_ISLEM_CHAT_CATEGORIES = " + chat_json + ";\n")
    payload_lines.append(
        "  window.normalizeTr = function(s){ return String(s||'').toLowerCase().replace(/\\u0131/g,'i').replace(/\\u011f/g,'g').replace(/\\u00fc/g,'u').replace(/\\u015f/g,'s').replace(/\\u00f6/g,'o').replace(/\\u00e7/g,'c').replace(/\\s+/g,' ').trim(); };\n"
    )
    payload_lines.append(
        "  window.getAraIslemRec = function(wasteType){ if(!wasteType) return null; var k = window.normalizeTr(wasteType); return window.ARA_ISLEM_RECOMMENDATIONS[k] || null; };\n"
    )
    payload_lines.append("})();\n")
    payload = "".join(payload_lines)

    for out_path in out_paths:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            f.write(payload)

    print("Generated", out_paths, "categories:", list(recommendations.keys()))


if __name__ == "__main__":
    main()

